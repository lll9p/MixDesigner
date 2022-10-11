#!/usr/bin/env python
# coding: utf-8
from collections import OrderedDict
from decimal import Decimal
from fractions import Fraction
from itertools import chain, zip_longest
from typing import Any, Dict, Iterable, List, Optional, Tuple, Type, TypeVar

import numpy as np
import openpyxl
import openpyxl.styles
import openpyxl.utils
import openpyxl.workbook.defined_name
from openpyxl.worksheet.table import Table, TableColumn, TableFormula
from typing_extensions import Self

from utils import convert_excel_array_attr_text, revert_excel_array_attr_text

from .BaseModel import BaseModel
from .types import (BoundsType, ComponentNamesType, DefaultsType, PointsType,
                    TargetNamesType, TestPointType)

ModelType = TypeVar("ModelType", bound="SimplexCentroid")

LIMIT_DENOMINATOR = 1000000


def np_formatter_func(x: Any) -> str:
    if isinstance(x, Fraction):
        return str(x.limit_denominator(LIMIT_DENOMINATOR))
    else:
        return str(x)


np.set_printoptions(formatter={"all": np_formatter_func})


class SimplexCentroid(BaseModel):
    """
    SimplexCentroid model for mixdesiner.
    Theory
    ------
    Convert proportion to projected space, fit the response surface coefficients.
    Experiments numbers equals to power(n_components,2)-1.
    Let n=n_components, en=2**n-1
    Powerset mask is (Removed first empty subset.)
    """

    name = "SimplexCentroid"
    parameters = [
        "name",
        "n_components",
        "component_names",
        "lower_bounds",
        "upper_bounds",
        "target_names",
        "n_experiments",
        "transform_matrix",
        "projected_matrix",
        "projected_matrix_test",
        "proportion",
        "proportion_test",
    ]

    def __init__(self, defaults: Optional[DefaultsType] = None) -> None:
        """
        Initial parameters for this estimator.

        Parameters
        ----------
        defaults : dict | None, default=None
            Specific the defaults of model.
            If it's a dict of params, use them to build model.
            If it's true of something not false or none, build with 3 components model,
            with 0 0 0 lb and 1 1 1 ub, with 3 test points.
            If it's none or false, build an empty model, then use `with_` methods to complete build.

        Returns
        -------
        None
        """
        super().__init__(name="SimplexCentroid")
        self.built = False
        if isinstance(defaults, dict):
            n_components = defaults.get("n_components")
            component_names = defaults.get("component_names")
            lower_bounds = defaults.get("lower_bounds")
            upper_bounds = defaults.get("upper_bounds")
            test_points = defaults.get("test_points")
            target_names = defaults.get("target_names")
            n_experiments = defaults.get("n_experiments")
            target_values = defaults.get("target_values")
            # fmt: off
            self.with_n_experiments(n=n_components)\
                .with_component_names(names=component_names)\
                .with_lower_bounds(bounds=lower_bounds)\
                .with_upper_bounds(bounds=upper_bounds)\
                .with_test_points(points=test_points)\
                .with_target_names(names=target_names)\
                .with_n_experiments(n=n_experiments)\
                .with_target_values(values=target_values)\
                .complete()
            # fmt: on
        elif defaults:
            self.complete()
        else:
            return

    @classmethod
    def build(cls: Type[ModelType]) -> ModelType:
        return cls()

    def with_n_components(self, n: Optional[int] = None) -> Self:
        if n is None:
            n = 3
        if n < 3:
            print(f"n_points must greater than 3, got {n}")
            return self
        self.n_components = n
        return self

    def with_component_names(
        self, names: Optional[ComponentNamesType] = None
    ) -> Self:
        if hasattr(self, "n_components"):
            if names is None:
                self.component_names = [
                    f"X{i}" for i in range(self.n_components)
                ]
            else:
                assert (
                    len(names) == self.n_components and len(names) >= 3
                ), "Number of components not match component names."
                self.component_names = names
        else:
            if names is None:
                # None of n_components and names provided. Choose defaults.
                print(
                    "Number of components and component names not found. "
                    "Choose defaults automatically (3)."
                )
                self.with_n_components(3).with_component_names()
            else:
                assert (
                    len(names) >= 3
                ), "Number of components names must greater than 3."
                self.with_n_components(len(names)).with_component_names(names)
        return self

    def with_lower_bounds(self, bounds: Optional[BoundsType] = None) -> Self:
        if not hasattr(self, "n_components"):
            print("Number of components not set.")
            return self
        self.lower_bounds = self._check_bounds(
            bounds=bounds, n_components=self.n_components, type_="lower"
        )
        return self

    def with_upper_bounds(self, bounds: Optional[BoundsType] = None) -> Self:
        if not hasattr(self, "n_components"):
            print("Number of components not set.")
            return self
        self.upper_bounds = self._check_bounds(
            bounds=bounds, n_components=self.n_components, type_="upper"
        )
        return self

    def with_test_points(self, points: Optional[TestPointType] = None) -> Self:
        if points is None:
            points = 3
        # **TODO** Check test points.
        self.test_points = points
        return self

    def with_target_names(
        self, names: Optional[TargetNamesType] = None
    ) -> Self:
        if names is None:
            names = ["Target0"]
        self.target_names = names
        return self

    def with_n_experiments(self, n: Optional[int] = None) -> Self:
        if n is None:
            n = 1
        self.n_experiments = n
        return self

    def with_target_values(
        self, values: Optional[Iterable[Iterable[float]]] = None
    ) -> Self:
        if values is None:
            self.target_values = None
        self.target_values = values
        return self

    def complete(self) -> Self:
        # If missing some params then make them default.
        if not hasattr(self, "n_components"):
            self.with_n_components()
        if not hasattr(self, "component_names"):
            self.with_component_names()
        if not hasattr(self, "lower_bounds"):
            self.with_lower_bounds()
        if not hasattr(self, "upper_bounds"):
            self.with_upper_bounds()
        if not hasattr(self, "test_points"):
            self.with_test_points()
        if not hasattr(self, "n_experiments"):
            self.with_n_experiments()
        if not hasattr(self, "target_names"):
            self.with_target_names()
        if not hasattr(self, "target_values"):
            self.with_target_values()

        self.experiment_points = self._generate_experiment_points()
        self.transform_matrix = self._generate_transform_matrix()
        # So called encoded proportion
        self.projected_matrix = self._generate_project_matrix()
        self.proportion = self._generate_proportion()
        # if test_points is not None:
        self.projected_matrix_test = self._generate_test_points(
            self.test_points
        )
        self.proportion_test = self.transform(self.projected_matrix_test)
        # So called real proportion
        self._response_surface_coefs: Optional[Dict[str, np.ndarray]] = None
        self.built = True
        return self

    @staticmethod
    def _check_bounds(
        bounds: Optional[BoundsType], n_components: int, type_: str
    ) -> np.ndarray[Fraction, Any]:
        make_fraction = lambda n: Fraction(str(n)).limit_denominator(
            LIMIT_DENOMINATOR
        )
        if bounds is None:
            bound = 0.0 if type_ == "lower" else 1.0
            bounds = tuple((make_fraction(bound),) * n_components)
        else:
            bounds = tuple(make_fraction(bound) for bound in bounds)
        return np.array(bounds)

    def _generate_proportion(self) -> PointsType:
        return self.transform(self.projected_matrix)

    def _generate_transform_matrix(self) -> np.ndarray[Fraction, Any]:
        # Get transform_matrix
        # Transform matrix
        lower_bounds = self.lower_bounds
        n = self.n_components
        # fmt: off
        transform_matrix = lower_bounds\
                .repeat(n)\
                .reshape((n, n))\
                + np.eye(n,dtype=Fraction)\
                * (1 - lower_bounds.sum())
        # fmt: on
        return transform_matrix

    def _generate_project_matrix(self) -> PointsType:
        """
        Take from experiment_points and divide component counts.
        """
        n = self.n_components
        experiment_points = self.experiment_points
        project_matrix = self.ndarray_to_fraction_array(
            np.ones((2**n - 1, n))
            * experiment_points
            / experiment_points.sum(axis=1)[:, np.newaxis]
        )
        return project_matrix

    def _generate_experiment_points(
        self,
    ) -> np.ndarray:
        """
        test_points:refer to proportion test_points.need to convert to nature points.
        """

        def bits(byte, n):
            """
            n-1 is total length of byte.
            for example, n=4, n-1=3, so,
            0b0001 -> 0001
            if use n, 0b0001 -> 00001
            """
            return [(byte & 1 << i) >> i for i in range(n - 1, -1, -1)]

        n = self.n_components
        # Generate a powerset mask except void set by start range from 1.
        experiment_points = np.array([bits(i, n=n) for i in range(1, 2**n)])
        return experiment_points

    def _generate_test_points(
        self, test_points: Optional[TestPointType] = 3
    ) -> PointsType:
        if isinstance(test_points, int):
            # random pick n_test_points does not need to transform_reverse
            added_points = np.random.dirichlet(
                (1,) * self.n_components, size=test_points
            )
        elif isinstance(test_points, List | Tuple | np.ndarray):
            # if nested
            added_points = np.array(test_points)
            # if not nested
            if not all(
                isinstance(i, List | Tuple | np.ndarray) for i in test_points
            ):
                # expand one dim
                added_points = np.expand_dims(added_points, axis=0)
            added_points = self.transform_reverse(added_points)
        else:
            added_points = np.random.dirichlet((1,) * self.n_components, size=3)
        return self.ndarray_to_fraction_array(added_points)

    def transform(self, points: PointsType) -> PointsType:
        """
        Transform projected points to proportion.
        """
        return points @ self.transform_matrix.T

    def transform_reverse(self, points: PointsType) -> PointsType:
        """
        Transform proportion to projected space.
        """
        inv_matrix = np.linalg.inv(self.transform_matrix.T.astype(np.float64))
        ret = points @ inv_matrix
        return ret

    def transform_variables(self, projected: np.ndarray):
        """
        Transform variable(x) to producted variables.
        Theory:
            A: the coefficients;
            X: the variables;
            then we got
            y= a_1x_1 + a_2x_2 + ... a_nx_1x_2..x_n           (e.q-1)
            X as variables is known; y as response value is known. We want to solve A.
            We have to transform X to x_1,x_2,..,x_n,x_1x_2,..,x_1x_2..x_n form.
            Just take x_n from experiment points(masks), then make a product.
            We can get A by solve a linear equation.

        Parameters
        ----------
        projected: np.ndarray

        Returns
        -------
        Producted X matrix
        """
        # Numpy use bool array as mask
        points = self.experiment_points.astype(bool)
        transpose = projected.T
        return np.array([transpose[point].prod(axis=0) for point in points]).T

    @staticmethod
    def ndarray_to_fraction_array(arr: PointsType) -> PointsType:
        number_to_fraction = lambda x: Fraction(
            Decimal(str(x))
        ).limit_denominator(LIMIT_DENOMINATOR)
        return np.vectorize(number_to_fraction)(arr)

    def fit(self, y):
        """
        generate the formula with specific y, y be experiment's results
        assume y's order same as model.test_points
        """
        if len(y) != len(self.experiment_points):
            raise TypeError(
                "Missing required positional argument: "
                "y's length not match test_points"
            )
        # Response surface coefs
        X = self.transform_variables(self.projected_matrix).astype(np.float64)
        return np.linalg.solve(X, y)

    def fit_targets(self) -> Self:
        """
        Average target values by target names. Then fit them.

        Parameters
        ----------
        None

        Returns
        -------
        None
        """
        # Average target values.
        if self.target_values is None:
            print("No target values set.")
            return self
        values = np.array(self.target_values)
        values_len = values.shape[0]
        n = self.n_experiments
        target_names = self.target_names
        # Group mean.
        averaged = np.array(
            [values[i : i + n].T.mean(axis=1) for i in range(0, values_len, n)]
        )
        y = averaged[:, : len(self.experiment_points)]
        coefs = {name: self.fit(y[i]) for i, name in enumerate(target_names)}
        self._response_surface_coefs = coefs
        return self

    def predict(self, proportion: Iterable, target_name: str):
        if self._response_surface_coefs is None:
            print("Not fit yet.")
            return
        proportion = np.array(proportion)
        # Get coefficients of some target.
        coef = self._response_surface_coefs[target_name]
        # Convert proportion to projected space.
        projected_proportion = self.transform_reverse(proportion)
        prediction = self.transform_variables(projected_proportion) @ coef
        return prediction

    def predict_targets(self) -> Optional[Dict[str, np.ndarray]]:
        if self._response_surface_coefs is None:
            print("Not fit yet.")
            return
        proportion = np.vstack([self.proportion, self.proportion_test])
        predictions = dict()
        for target_name in self.target_names:
            predictions[target_name] = self.predict(proportion, target_name)
        return predictions

    def predict_samples(
        self, proportion: np.ndarray
    ) -> Optional[Dict[str, np.ndarray]]:
        if self._response_surface_coefs is None:
            print("Not fit yet.")
            return
        predictions = dict()
        for target_name in self.target_names:
            predictions[target_name] = self.predict(proportion, target_name)
        return predictions

    def save_to_file(self, file: str) -> None:
        workbook = openpyxl.Workbook()
        # Writting conditions.
        make_sheet_conditions(workbook=workbook, model=self)
        # Writting experiments
        make_sheet_experiments(workbook=workbook, model=self)
        # Writting coefficients
        make_sheet_coefficients(workbook=workbook, model=self)
        # Writting calculations.
        make_sheet_calculations(workbook=workbook, model=self)
        # Remove default sheet
        if "Sheet" in workbook.sheetnames:
            workbook.remove_sheet(workbook["Sheet"])
        try:
            workbook.save(file)
            workbook.close()
        except Exception as e:
            print(e)

    @classmethod
    def build_from_file(cls: Type[ModelType], file: str) -> Optional[ModelType]:
        """
        Build model from exsit excel file.
        """
        workbook = openpyxl.load_workbook(file, read_only=True, data_only=True)
        # Get conditions from workbook defined names.
        (
            model_name,
            component_names,
            n_components,
            lower_bounds,
            upper_bounds,
            n_experiments,
        ) = read_conditions(workbook=workbook)
        if model_name != cls.name:
            print(
                f"Your excel file seems not for model:{cls.name}. Model name in excel file is {model_name}."
            )
            return
        target_names, target_values, test_points = read_experiments(
            workbook=workbook,
            n_components=n_components,
            n_experiments=n_experiments,
        )
        workbook.close()
        model = (
            cls.build()
            .with_n_components(n=n_components)
            .with_component_names(names=component_names)
            .with_lower_bounds(bounds=lower_bounds)
            .with_upper_bounds(bounds=upper_bounds)
            .with_test_points(points=test_points)
            .with_target_names(names=target_names)
            .with_n_experiments(n=n_experiments)
            .with_target_values(values=target_values)
            .complete()
        )
        return model

    def get_params(self, param: Optional[str] = None):
        def get(obj: str) -> Any:
            if hasattr(self, obj):
                return getattr(self, obj)
            else:
                return "Not set yet."

        if not self.built:
            return "Model not built."
        if param:
            return f"{param}:\n{get(param)}"
        return f"""
ModelName:\n{get("name")}\n\n
NComponents:\n{get("n_components")}\n
Components:\n{get("component_names")}\n
LowerBounds:\n{get("lower_bounds")}\n
UpperBounds:\n{get("upper_bounds")}\n
Targets:\n{get("target_names")}\n
NExperiments:\n{get("n_experiments")}\n
TransformMatrix:\n{get("transform_matrix")}\n
ProjectedMatrix:\n{get("projected_matrix")}\n
ProjectedMatrixTest:\n{get("projected_matrix_test")}\n
Proportion:\n{get("proportion")}\n
ProportionTest:\n{get("proportion_test")}
"""


def make_sheet_conditions(workbook: openpyxl.Workbook, model: SimplexCentroid):
    """
    Make 'Conditions' sheet.
    """

    def make_defined_array(
        name: str, source: List[List[str]], workbook: openpyxl.Workbook
    ):
        attr_text = convert_excel_array_attr_text(source)
        defined_name = openpyxl.workbook.defined_name.DefinedName
        defined_array = defined_name(name=name, attr_text=attr_text)
        workbook.defined_names.append(defined_array)

    sheet = workbook.create_sheet("Conditions")
    # name arr expand
    arrays = (
        ("model_name", [[model.name]], 1),
        (
            "component_names",
            [model.component_names],
            len(model.component_names),
        ),
        (
            "lower_bounds",
            [model.lower_bounds.astype("str").tolist()],
            model.lower_bounds.shape[0],
        ),
        (
            "upper_bounds",
            [model.upper_bounds.astype("str").tolist()],
            model.upper_bounds.shape[0],
        ),
        ("n_experiments", [[str(model.n_experiments)]], 1),
    )
    for i, (name, arr, expand) in enumerate(arrays):
        make_defined_array(name, arr, workbook)
        func = openpyxl.utils.get_column_letter
        start_letter = func(1 + 1)
        end_letter = func(expand + 1)
        coord = f"{start_letter}{i+1}"
        coord_range = f"{coord}:{end_letter}{i+1}"
        sheet.formula_attributes[coord] = dict(t="array", ref=coord_range)
    data = [
        ["Model", "=model_name"],
        ["Names", "=component_names"],
        ["Lower bounds", "=lower_bounds"],
        ["Upper bounds", "=upper_bounds"],
        ["Experiments numbers", "=n_experiments"],
    ]
    for line in data:
        sheet.append(line)


def make_sheet_experiments(workbook: openpyxl.Workbook, model: SimplexCentroid):
    """
    Make 'Experiments' sheet.
    """
    sheet = workbook.create_sheet("Experiments")
    # Writting header
    header = (
        ["№"]
        + model.component_names
        + list(
            chain.from_iterable(
                ([name] * model.n_experiments for name in model.target_names)
            )
        )
    )
    sheet.append(header)
    # Proportion with test points.
    proportion = np.vstack((model.proportion, model.proportion_test))
    if model.target_values is not None:
        target_values = model.target_values
    else:
        target_values = (
            [[None] * proportion.shape[0]]
            * len(model.target_names)
            * model.n_experiments
        )
    lines = (
        np.hstack((proportion, np.array(target_values).T))
        .astype(np.float64)
        .tolist()
    )
    test_id = 0
    for line, point in zip_longest(lines, model.experiment_points.tolist()):
        if point is None:
            exp_id = f"TEST-{test_id}"
            test_id += 1
        else:
            exp_id = "EXP-" + "".join(str(s) for s in point)
        sheet.append([exp_id] + line)


def make_sheet_coefficients(
    workbook: openpyxl.Workbook, model: SimplexCentroid
):
    """
    Make 'Coefficients' sheet.
    """
    sheet = workbook.create_sheet("Coefficients")
    # Save coefs
    if model.target_values is not None:
        model.fit_targets()
    target_names_array = np.array(model.component_names)
    variables = [
        "*".join(target_names_array[ei])
        for ei in model.experiment_points.astype(bool)
    ]

    coefs_with_header = [["Variables"] + variables]
    for name in model.target_names:
        if model._response_surface_coefs is not None:
            coef = model._response_surface_coefs.get(name)
            if isinstance(coef, np.ndarray):
                coef = coef.tolist()
            else:
                coef = []
        else:
            coef = []
        coefs = [name] + coef
        coefs_with_header.append(coefs)
    # Writting
    for coef in zip(*coefs_with_header):
        sheet.append(coef)

    # Make a table for coefficients.
    table_columns = []
    for i, col in enumerate(coefs_with_header, start=1):
        table_columns.append(TableColumn(id=i, name=col[0]))
    col_name = openpyxl.utils.get_column_letter(len(coefs_with_header))
    tab = Table(
        displayName="Coefficients",
        ref=f"A1:{col_name}{len(coefs_with_header[0])}",
    )
    sheet.add_table(tab)


def make_sheet_calculations(
    workbook: openpyxl.Workbook, model: SimplexCentroid
):
    """
    Make 'Calculations' sheet.
    """
    sheet_calc = workbook.create_sheet("Calculations")
    # Create calculation table data.
    target_names_array = np.array(model.component_names)
    variables = [
        "*".join(target_names_array[ei])
        for ei in model.experiment_points.astype(bool)
    ]
    table_header = (
        model.component_names  # Components
        + model.target_names  # Targets
        + [name + "_" for name in model.component_names]  # Projected components
        + ["(" + variable + ")" for variable in variables]
    )
    sheet_calc.append(table_header)
    # Add transform matrix to defined_array
    transform_matrix = convert_excel_array_attr_text(
        model.transform_matrix.astype(np.float64).astype(str).tolist(),
        number=True,
    )
    defined_name = openpyxl.workbook.defined_name.DefinedName
    defined_array = defined_name(
        name="transform_matrix", attr_text=transform_matrix
    )
    workbook.defined_names.append(defined_array)

    target_value_format_string = (
        "=SUM(TRANSPOSE(Coefficients[{}])*Calculations[[#This Row],[{}]:[{}]])"
    )
    target_values_formulaes = []
    for target_name in model.target_names:
        formulaes = target_value_format_string.format(
            target_name, f"({variables[0]})", f"({variables[-1]})"
        )
        target_values_formulaes.append(formulaes)
    projected_format_string = "=INDEX(MMULT(Calculations[[#This Row],[{}]:[{}]],MINVERSE(TRANSPOSE({}))),{})"
    projected_formulaes = []
    for i in range(1, model.n_components + 1):
        formulaes = projected_format_string.format(
            model.component_names[0],
            model.component_names[-1],
            "transform_matrix",
            i,
        )
        projected_formulaes.append(formulaes)
    variables_formulaes = []
    for ei in model.experiment_points.astype(bool):
        arr = target_names_array[ei]
        v = "=" + "*".join([f"[[#This Row],[{s}_]]" for s in arr])
        variables_formulaes.append(v)
    table_col_formulaes = (
        target_values_formulaes + projected_formulaes + variables_formulaes
    )
    proportion = np.vstack((model.proportion, model.proportion_test))
    for line in proportion.astype(np.float64).tolist():
        sheet_calc.append(line + table_col_formulaes)
    table_data_cols = (
        model.proportion.shape[1]
        + len(target_values_formulaes)
        + len(projected_formulaes)
        + len(variables_formulaes)
    )
    table_data_rows = proportion.shape[0]
    for row_index in range(2, table_data_rows + 2):
        for col_index in range(1, table_data_cols + 1):
            col_letter = openpyxl.utils.get_column_letter(col_index)
            cell_coord = f"{col_letter}{row_index}"
            sheet_calc.formula_attributes[cell_coord] = {
                "t": "array",
                "ref": cell_coord,
            }
    table_header_col_letter = openpyxl.utils.get_column_letter(
        len(table_header)
    )
    table_row_num = table_data_rows + 1
    table_columns = []
    for i in range(1, table_data_cols + 1):
        if i <= model.proportion.shape[1]:
            tmp = TableColumn(id=i, name=table_header[i - 1])
        else:
            tmp = TableColumn(
                id=i,
                name=table_header[i - 1],
                calculatedColumnFormula=TableFormula(
                    array=True,
                    attr_text=table_col_formulaes[
                        i - model.proportion.shape[1] - 1
                    ],
                ),
            )
        table_columns.append(tmp)

    tab = Table(
        displayName="Calculations",
        ref=f"A1:{table_header_col_letter}{table_row_num}",
        tableColumns=table_columns,
    )
    sheet_calc.add_table(tab)


def read_conditions(
    workbook: openpyxl.Workbook,
) -> Tuple[str, List[str], int, BoundsType, BoundsType, int]:
    """
    Read conditions
    """

    def get_array(name: str) -> List[List[str]]:
        return revert_excel_array_attr_text(
            workbook.defined_names.get(name).attr_text
        )

    model_name = get_array("model_name")[0][0]
    component_names = get_array("component_names")[0]
    n_components = len(component_names)
    lower_bounds = list(map(Fraction, get_array("lower_bounds")[0]))
    upper_bounds = list(map(Fraction, get_array("upper_bounds")[0]))
    n_experiments = int(get_array("n_experiments")[0][0])
    return (
        model_name,
        component_names,
        n_components,
        lower_bounds,
        upper_bounds,
        n_experiments,
    )


def read_experiments(
    workbook: openpyxl.Workbook, n_components: int, n_experiments: int
) -> Tuple[List[str], np.ndarray, np.ndarray]:
    """
    Read experiments from workbook.
    """
    sheet = workbook["Experiments"]
    values = np.array(tuple(sheet.values))
    header = values[0]
    header_target_names = header[n_components + 1 :]
    if len(header_target_names) != n_components * n_experiments:
        print("Your experiment is not complete yet. Will ignore target values.")
    target_names = list(OrderedDict.fromkeys(header_target_names))
    target_values = values[1:, n_components + 1 :].T.astype(np.float64)
    test_points = values[
        np.where(np.char.startswith(values[:, 0], "TEST-"))[0],
        1 : n_components + 1,
    ].astype(np.float64)
    return target_names, target_values, test_points
